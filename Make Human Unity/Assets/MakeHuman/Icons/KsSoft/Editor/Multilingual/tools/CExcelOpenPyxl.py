#!/usr/bin/env python
# -*- coding: utf-8 -*-
from openpyxl import *

import sys
import os.path
import re
import warnings
from CVariable import CReadVariable,CWriteVariable
from CError import *

class CExcelOpenPyxl:
	#==============================================
	#!Constructor
	# @brief constructor
	#==============================================
	def __init__(self,xlsfile):
		self.name = xlsfile
		warnings.simplefilter("ignore")
		self.workbook = reader.excel.load_workbook(xlsfile)
		self.sheetnames = self.workbook.get_sheet_names()
	#==============================================
	#!getSheetNum
	# @brief シートの総数を取得
	#==============================================
	def	getSheetNum(self):
		return len(self.sheetnames)
	#==============================================
	#!getSheetName
	# @brief シートの名前を取得
	#==============================================
	def getSheetName(self,index):
		if (index >= len(self.sheetnames)):
			return None
		return self.sheetnames[index]
	#==============================================
	#!getSheet
	# @brief シートを取得
	#==============================================
	def getSheet(self,index):
		if (index >= self.getSheetNum()):
			return None
		return self.workbook.worksheets[index]
	#==============================================
	#!getSheetByName
	# @brief シートを名前を使って取得
	#==============================================
	def getSheetByName(self,sheetname):
		return self.workbook.get_sheet_by_name(sheetname)
	#==============================================
	#!convert
	# @brief シートのデータを連想配列の配列形式に変更する
	#==============================================
	def convert(self,sheet,config = 'D'):
		height = sheet.max_row + 1
		width = sheet.max_column + 1
		rsel = re.compile('\[' + config + '\]')
		tags = []
		# 原点[ORIGIN]を探す
		ox = None
		oy = None
		row_sel = None
		for y in range(height):
			for x in range(width):
				if (sheet.cell(row=y, column=x).value == '[ORIGIN]'):
					ox = x
					oy = y
					# Debug,Releaseをを選択
					for x in range(ox + 1,width):
						v = sheet.cell(row=oy + 1,column=x).value
						if (v is not None and rsel.search(v) is not None):
							tag = sheet.cell(row=oy,column=x).value
							if (tag is None):
								# 行単位で選択するためのカラム
								row_sel = x
							else:
								tags.append([tag,x])
					break
			if (ox is not None):
				break
		if (oy is None):
#			print "can't find keyword '[ORIGIN]'"
			return None
		# 結果を連想配列にして突っ込む
		r = []
		for y in range(oy + 2,height):
			one = {}
			skip = False
			for tag in tags:
				if (row_sel is not None and sheet.cell(row=y,column=row_sel).value is None):
					skip = True
					break
				cell = sheet.cell(row=y,column=tag[1]);
				v = cell.value
				one[tag[0]] = v;
			if (not skip):
				r.append(one)
		return r
	#==============================================
	#!create
	# @brief シートデータを特定の値をキーとした連想配列に格納する
	#==============================================
	def create(self,mapSheet,sheetname,keytag,cError,config = 'D'):
		sheet = self.getSheetByName(sheetname)
		line = self.convert(sheet,config)
		nLine = 0
		result = True
		for one in line:
			if (keytag not in one):
				cError.addPlain("can't find key:" + keytag,self.name + ":" + sheet,nLine)
				return False
			key = one[keytag]
			if (key is None):
				continue
			if (key in mapSheet):
				cError.addPlain("already exist key:" + key,self.name + ":" + sheet,nLine)
				result = False
			mapSheet[key] = one
			nLine += 1
		return result
	#==============================================
	#!create
	# @brief シートデータを特定の値をキーとした連想配列に格納する
	#==============================================
	def createValueMap(self,mapSheet,sheetname,keytag,valuetag,cError,config = 'D'):
		sheet = self.getSheetByName(sheetname)
		line = self.convert(sheet,config)
		nLine = 0
		result = True
		for one in line:
			if (keytag not in one):
				cError.addPlain("can't find key tag:" + keytag,self.name + ":" + sheet,nLine)
				return False
			if (valuetag not in one):
				cError.addPlain("can't find value tag:" + valuetag,self.name + ":" + sheet,nLine)
				return False
			key = one[keytag]
			if (key is None):
				continue
			if (key in mapSheet):
				cError.addPlain("already exist key:" + key,self.name + ":" + sheet,nLine)
				result = False
			mapSheet[key] = one[valuetag]
			nLine += 1
		return result
	#==============================================
	#!create
	# @brief シートデータを特定の値をキーとした連想配列に格納する
	#==============================================
	def createMulIdMap(self,mapSheet,sheetname,keytag,valuetag,cError,config = 'D'):
		sheet = self.getSheetByName(sheetname)
		line = self.convert(sheet,config)
		nLine = 0
		result = True
		for one in line:
			if (keytag not in one):
				cError.addPlain("can't find key tag:" + keytag,self.name + ":" + sheet,nLine)
				return False
			if (valuetag not in one):
				cError.addPlain("can't find value tag:" + valuetag,self.name + ":" + sheet,nLine)
				return False
			key = one[keytag]
			if (key is None):
				continue
			if (key in mapSheet):
				cError.addPlain("already exist key:" + key,self.name + ":" + sheet,nLine)
				result = False
			val = CMulId(one[valuetag],cError).value
			mapSheet[key] = val
			mapSheet[one[valuetag]] = val
			nLine += 1
		return result		
	#==============================================
	#!getFromMap
	# @brief 連想配列化したシートから特定の値を取り出す
	#==============================================
	def getFromMap(self,mapSheet,key,cError):
		if (key is None):
			return 0
		if (key not in mapSheet):
			cError.add("can't find value " + key)
			return None
		return mapSheet[key]

	#
	# 以下のコマンドライン引数を解析する
	#   cmd [-v version] excelfile [outputfile]
	#  param:excelpath	:エクセルパス
	#  param:version	:Excelで有効にするカラムを選択
	#  param:help		:エラー発生時に呼び出すヘルプ
	#
	@staticmethod
	def checkArgv(excelpath,version,help):
		#初期値
		srcpath = None
		outpath = None
		# 引数解析
		narg = len(sys.argv)
		i = 1
		while (i < narg):
			arg = sys.argv[i]
			if (arg == '-v'):
				i += 1
				if (i >= narg):
					help()
					sys.exit()
				version = sys.argv[i]
			elif (arg == '-h'):
				help()
				sys.exit()
			else:
				if (srcpath == None):
					srcpath = arg
				elif (outpath == None):
					outpath = arg
				else:
					help()
					sys.exit()
			i += 1
		if (srcpath == None):
			srcpath = excelpath
		if (outpath == None):
			outpath,ext = os.path.splitext(srcpath)
		return [srcpath,outpath,version]
